"""
╔══════════════════════════════════════════════════════════════════╗
║           ETL DATAWAREHOUSE — TP Activité Commerciale           ║
║  Sources : MySQL SQL | TXT | Excel | JSON                       ║
║  Destination : PostgreSQL (schéma en étoile)                    ║
╚══════════════════════════════════════════════════════════════════╝

Architecture cible :
  FACT_SALES_ACTIVITY
    ← DIM_DATE
    ← DIM_SELLER
    ← DIM_CUSTOMER
    ← DIM_PRODUCT
    ← DIM_GEO

Usage :
  python3 etl_pipeline.py

Sortie :
  05_load_postgres.sql
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path


# ─── Config chemins ───────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
SQL_FILE = BASE_DIR / "01_source_mysql_printer_sales.sql"
TXT_FILE = BASE_DIR / "02_route_logs.txt"
EXCEL_FILE = BASE_DIR / "03_sales_promises.xlsx"
JSON_FILE = BASE_DIR / "04_fuel_expenses.json"
OUT_SQL = BASE_DIR / "05_load_postgres.sql"


# ─── Couleurs terminal ────────────────────────────────────────────────────────
G = "\033[92m"
Y = "\033[93m"
R = "\033[91m"
B = "\033[94m"
E = "\033[0m"


def ok(msg):
    print(f"  {G}✓{E} {msg}")


def info(msg):
    print(f"  {B}→{E} {msg}")


def warn(msg):
    print(f"  {Y}⚠{E} {msg}")


def err(msg):
    print(f"  {R}✗{E} {msg}")


# ══════════════════════════════════════════════════════════════════════════════
# UTILITAIRES NORMALISATION
# ══════════════════════════════════════════════════════════════════════════════
def clean_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_code(value) -> str:
    return clean_str(value).upper()


def norm_city(value) -> str:
    raw = clean_str(value)
    if not raw:
        return "INCONNUE"
    return raw.upper()


def norm_status_order(value) -> str:
    status = clean_str(value).upper()
    mapping = {
        "DELIVERED": "DELIVERED",
        "INVOICED": "INVOICED",
        "CONFIRMED": "CONFIRMED",
        "CANCELLED": "CANCELLED",
    }
    return mapping.get(status, "OTHER")


def norm_status_promise(value) -> str:
    status = clean_str(value).upper()
    mapping = {
        "COLD": "COLD",
        "WARM": "WARM",
        "HOT": "HOT",
        "WON": "WON",
        "LOST": "LOST",
    }
    return mapping.get(status, "OTHER")


def parse_decimal(value, default=0.0) -> float:
    if value is None:
        return float(default)
    if isinstance(value, (int, float)):
        return float(value)

    txt = clean_str(value)
    if txt == "":
        return float(default)

    txt = txt.replace(" ", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return float(default)


def parse_int(value, default=0) -> int:
    if value is None:
        return int(default)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))

    txt = clean_str(value)
    if txt == "":
        return int(default)

    txt = txt.replace(" ", "")
    if "," in txt:
        return int(round(parse_decimal(txt, default=default)))

    try:
        return int(txt)
    except ValueError:
        try:
            return int(round(float(txt)))
        except ValueError:
            return int(default)


def parse_date(value) -> str:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    txt = clean_str(value)
    if txt == "":
        return None

    formats = ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"]
    for fmt in formats:
        try:
            return datetime.strptime(txt, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def esc_sql(value: str) -> str:
    return value.replace("'", "''")


# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 1 — EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def split_sql_values(values_str: str) -> list[str]:
    out = []
    token = []
    in_quote = False
    i = 0
    n = len(values_str)

    while i < n:
        ch = values_str[i]
        if ch == "'":
            if in_quote and i + 1 < n and values_str[i + 1] == "'":
                token.append("'")
                i += 2
                continue
            in_quote = not in_quote
            i += 1
            continue

        if ch == "," and not in_quote:
            out.append("".join(token).strip())
            token = []
            i += 1
            continue

        token.append(ch)
        i += 1

    if token:
        out.append("".join(token).strip())

    return out


def extract_mysql_sql(path: Path):
    """Parse le dump MySQL et extrait vendeurs, clients, produits, commandes."""
    sellers = []
    customers = []
    products = []
    orders = []

    pattern = re.compile(r"INSERT INTO\s+(\w+)\s+\([^)]*\)\s+VALUES\s*\((.*)\);$")

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line.startswith("INSERT INTO"):
            continue

        m = pattern.match(line)
        if not m:
            continue

        table = m.group(1)
        values = split_sql_values(m.group(2))

        if table == "hr_vendeurs" and len(values) == 10:
            sellers.append(
                {
                    "seller_code": norm_code(values[0]),
                    "first_name": clean_str(values[1]).title(),
                    "last_name": clean_str(values[2]).upper(),
                    "email": clean_str(values[3]).lower(),
                    "salary": parse_decimal(values[4]),
                    "hire_date": parse_date(values[5]),
                    "home_country": clean_str(values[6]),
                    "home_region": clean_str(values[7]),
                    "home_city": clean_str(values[8]),
                    "manager_code": norm_code(values[9]),
                }
            )

        elif table == "ref_clients" and len(values) == 8:
            customers.append(
                {
                    "customer_code": norm_code(values[0]),
                    "customer_name": clean_str(values[1]),
                    "sector": clean_str(values[2]),
                    "country": clean_str(values[3]),
                    "city": clean_str(values[4]),
                    "postal_code": clean_str(values[5]),
                    "created_at": parse_date(values[6]),
                    "region_name": clean_str(values[7]),
                }
            )

        elif table == "ref_produits" and len(values) == 7:
            products.append(
                {
                    "product_code": norm_code(values[0]),
                    "product_name": clean_str(values[1]),
                    "range_name": clean_str(values[2]),
                    "list_price": parse_decimal(values[3]),
                    "active_flag": parse_int(values[4]),
                    "launch_date": parse_date(values[5]),
                    "category_name": clean_str(values[6]),
                }
            )

        elif table == "sales_orders" and len(values) == 11:
            orders.append(
                {
                    "order_id": parse_int(values[0]),
                    "order_date": parse_date(values[1]),
                    "seller_code": norm_code(values[2]),
                    "customer_code": norm_code(values[3]),
                    "product_code": norm_code(values[4]),
                    "quantity": parse_int(values[5]),
                    "unit_price": parse_decimal(values[6]),
                    "discount_pct": parse_decimal(values[7]),
                    "order_status": norm_status_order(values[8]),
                    "promised_delivery_date": parse_date(values[9]),
                    "created_at": parse_date(values[10]),
                }
            )

    return sellers, customers, products, orders


def extract_route_logs(path: Path) -> list:
    """Parse le fichier TXT route_logs (séparateur |)."""
    routes = []
    with path.open("r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="|")
        for row in reader:
            routes.append(
                {
                    "route_id": clean_str(row.get("route_id")),
                    "visit_date": parse_date(row.get("visit_date")),
                    "seller_code": norm_code(row.get("seller_code")),
                    "customer_code": norm_code(row.get("customer_code")),
                    "visit_city": norm_city(row.get("visit_city")),
                    "planned_visits": parse_int(row.get("planned_visits")),
                    "actual_visits": parse_int(row.get("actual_visits")),
                    "km_travelled": parse_decimal(row.get("km_travelled")),
                    "travel_expense": parse_decimal(row.get("travel_expense")),
                    "road_toll": parse_decimal(row.get("road_toll")),
                    "trip_status": clean_str(row.get("trip_status")).upper(),
                    "notes": clean_str(row.get("notes")),
                }
            )
    return routes


def extract_excel_promises(path: Path) -> list:
    """Lit le fichier Excel des promesses de vente."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    promises = []
    for row in range(2, ws.max_row + 1):
        vals = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        if all(v is None for v in vals):
            continue
        data = dict(zip(headers, vals))
        promises.append(
            {
                "promise_id": clean_str(data.get("promise_id")),
                "promise_date": parse_date(data.get("promise_date")),
                "seller_code": norm_code(data.get("seller_code")),
                "customer_code": norm_code(data.get("customer_code")),
                "product_code": norm_code(data.get("product_code")),
                "promised_qty": parse_int(data.get("promised_qty")),
                "expected_amount": parse_decimal(data.get("expected_amount"), default=None)
                if data.get("expected_amount") is not None
                else None,
                "expected_closing_date": parse_date(data.get("expected_closing_date")),
                "probability_pct": parse_decimal(data.get("probability_pct")),
                "status": norm_status_promise(data.get("status")),
                "sales_stage": clean_str(data.get("sales_stage")),
            }
        )
    return promises


def extract_fuel_json(path: Path) -> list:
    """Charge le JSON des frais carburant / déplacement."""
    rows = json.loads(path.read_text(encoding="utf-8"))
    fuel = []
    for r in rows:
        fuel.append(
            {
                "expense_id": clean_str(r.get("expense_id")),
                "seller_code": norm_code(r.get("seller_code")),
                "expense_date": parse_date(r.get("expense_date")),
                "fuel_liters": parse_decimal(r.get("fuel_liters")),
                "fuel_unit_price": parse_decimal(r.get("fuel_unit_price")),
                "fuel_cost": parse_decimal(r.get("fuel_cost")),
                "toll_cost": parse_decimal(r.get("toll_cost")),
                "hotel_cost": parse_decimal(r.get("hotel_cost")),
                "meal_cost": parse_decimal(r.get("meal_cost")),
                "misc_cost": parse_decimal(r.get("misc_cost")),
                "payment_mode": clean_str(r.get("payment_mode")).upper(),
                "receipt_status": clean_str(r.get("receipt_status")).upper(),
                "region_hint": clean_str(r.get("region_hint")),
            }
        )
    return fuel


# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 2 — TRANSFORMATION
# ══════════════════════════════════════════════════════════════════════════════
def transform_dim_date(orders: list, promises: list, routes: list, fuel: list) -> list:
    dates = set()
    for o in orders:
        if o["order_date"]:
            dates.add(o["order_date"])
        if o["promised_delivery_date"]:
            dates.add(o["promised_delivery_date"])
    for p in promises:
        if p["promise_date"]:
            dates.add(p["promise_date"])
        if p["expected_closing_date"]:
            dates.add(p["expected_closing_date"])
    for r in routes:
        if r["visit_date"]:
            dates.add(r["visit_date"])
    for f in fuel:
        if f["expense_date"]:
            dates.add(f["expense_date"])

    month_names = [
        "Janvier",
        "Fevrier",
        "Mars",
        "Avril",
        "Mai",
        "Juin",
        "Juillet",
        "Aout",
        "Septembre",
        "Octobre",
        "Novembre",
        "Decembre",
    ]
    day_names = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

    dim = []
    for i, d_str in enumerate(sorted(dates), 1):
        d = datetime.strptime(d_str, "%Y-%m-%d").date()
        dim.append(
            {
                "id_dim_date": i,
                "date_complete": d_str,
                "annee": d.year,
                "trimestre": ((d.month - 1) // 3) + 1,
                "mois": d.month,
                "lib_mois": month_names[d.month - 1],
                "semaine": d.isocalendar()[1],
                "jour": d.day,
                "lib_jour": day_names[d.weekday()],
                "est_weekend": d.weekday() >= 5,
            }
        )
    return dim


def transform_dim_seller(sellers: list, orders: list, promises: list, routes: list, fuel: list) -> list:
    by_code = {}

    for s in sellers:
        code = norm_code(s["seller_code"])
        by_code[code] = {
            "seller_code": code,
            "first_name": clean_str(s["first_name"]).title(),
            "last_name": clean_str(s["last_name"]).upper(),
            "email": clean_str(s["email"]).lower(),
            "salary": parse_decimal(s["salary"]),
            "hire_date": s["hire_date"],
            "home_country": clean_str(s["home_country"]),
            "home_region": clean_str(s["home_region"]),
            "home_city": clean_str(s["home_city"]),
            "manager_code": norm_code(s["manager_code"]),
        }

    all_codes = set()
    for src in (orders, promises, routes, fuel):
        for row in src:
            all_codes.add(norm_code(row.get("seller_code")))

    for code in sorted(all_codes):
        if not code:
            continue
        if code not in by_code:
            by_code[code] = {
                "seller_code": code,
                "first_name": "Inconnu",
                "last_name": "INCONNU",
                "email": None,
                "salary": 0.0,
                "hire_date": None,
                "home_country": None,
                "home_region": None,
                "home_city": None,
                "manager_code": None,
            }

    dim = []
    for i, code in enumerate(sorted(by_code.keys()), 1):
        r = by_code[code]
        full_name = f"{r['first_name']} {r['last_name']}".strip()
        dim.append(
            {
                "id_dim_seller": i,
                "seller_code": code,
                "first_name": r["first_name"],
                "last_name": r["last_name"],
                "full_name": full_name,
                "email": r["email"],
                "salary": r["salary"],
                "hire_date": r["hire_date"],
                "home_country": r["home_country"],
                "home_region": r["home_region"],
                "home_city": r["home_city"],
                "manager_code": r["manager_code"],
            }
        )
    return dim


def transform_dim_customer(customers: list, orders: list, promises: list, routes: list) -> list:
    by_code = {}

    for c in customers:
        code = norm_code(c["customer_code"])
        by_code[code] = {
            "customer_code": code,
            "customer_name": clean_str(c["customer_name"]),
            "sector": clean_str(c["sector"]),
            "country": clean_str(c["country"]),
            "region_name": clean_str(c["region_name"]),
            "city": clean_str(c["city"]),
            "postal_code": clean_str(c["postal_code"]),
            "created_at": c["created_at"],
        }

    all_codes = set()
    for src in (orders, promises, routes):
        for row in src:
            all_codes.add(norm_code(row.get("customer_code")))

    for code in sorted(all_codes):
        if not code:
            continue
        if code not in by_code:
            by_code[code] = {
                "customer_code": code,
                "customer_name": f"Client {code}",
                "sector": "Inconnu",
                "country": None,
                "region_name": None,
                "city": None,
                "postal_code": None,
                "created_at": None,
            }

    dim = []
    for i, code in enumerate(sorted(by_code.keys()), 1):
        r = by_code[code]
        dim.append(
            {
                "id_dim_customer": i,
                "customer_code": code,
                "customer_name": r["customer_name"],
                "sector": r["sector"],
                "country": r["country"],
                "region_name": r["region_name"],
                "city": r["city"],
                "postal_code": r["postal_code"],
                "created_at": r["created_at"],
            }
        )
    return dim


def transform_dim_product(products: list, orders: list, promises: list) -> list:
    by_code = {}

    for p in products:
        code = norm_code(p["product_code"])
        by_code[code] = {
            "product_code": code,
            "product_name": clean_str(p["product_name"]),
            "category_name": clean_str(p["category_name"]),
            "range_name": clean_str(p["range_name"]),
            "list_price": parse_decimal(p["list_price"]),
            "active_flag": parse_int(p["active_flag"]),
            "launch_date": p["launch_date"],
        }

    all_codes = set()
    for src in (orders, promises):
        for row in src:
            all_codes.add(norm_code(row.get("product_code")))

    for code in sorted(all_codes):
        if not code:
            continue
        if code not in by_code:
            by_code[code] = {
                "product_code": code,
                "product_name": f"Produit {code}",
                "category_name": "Inconnue",
                "range_name": "Inconnue",
                "list_price": 0.0,
                "active_flag": 0,
                "launch_date": None,
            }

    dim = []
    for i, code in enumerate(sorted(by_code.keys()), 1):
        r = by_code[code]
        dim.append(
            {
                "id_dim_product": i,
                "product_code": code,
                "product_name": r["product_name"],
                "category_name": r["category_name"],
                "range_name": r["range_name"],
                "list_price": r["list_price"],
                "active_flag": bool(r["active_flag"]),
                "launch_date": r["launch_date"],
            }
        )
    return dim


def transform_dim_geo(customers: list, routes: list, fuel: list, sellers: list) -> list:
    geo_map = {}

    for c in customers:
        city = norm_city(c.get("city"))
        if city not in geo_map:
            geo_map[city] = {
                "city": city,
                "region_name": clean_str(c.get("region_name")) or clean_str(c.get("country")),
                "country": clean_str(c.get("country")),
            }

    for r in routes:
        city = norm_city(r.get("visit_city"))
        if city not in geo_map:
            geo_map[city] = {
                "city": city,
                "region_name": None,
                "country": None,
            }

    for f in fuel:
        city = norm_city(f.get("region_hint"))
        if city not in geo_map:
            geo_map[city] = {
                "city": city,
                "region_name": clean_str(f.get("region_hint")),
                "country": None,
            }

    for s in sellers:
        city = norm_city(s.get("home_city"))
        if city not in geo_map:
            geo_map[city] = {
                "city": city,
                "region_name": clean_str(s.get("home_region")),
                "country": clean_str(s.get("home_country")),
            }

    if "INCONNUE" not in geo_map:
        geo_map["INCONNUE"] = {
            "city": "INCONNUE",
            "region_name": "Inconnue",
            "country": "Inconnu",
        }

    dim = []
    for i, city in enumerate(sorted(geo_map.keys()), 1):
        r = geo_map[city]
        dim.append(
            {
                "id_dim_geo": i,
                "city": r["city"],
                "region_name": r["region_name"] or "Inconnue",
                "country": r["country"] or "Inconnu",
            }
        )
    return dim


def transform_promises(promises: list, product_price_index: dict) -> list:
    out = []
    for p in promises:
        expected_amount = p["expected_amount"]
        if expected_amount is None:
            list_price = product_price_index.get(p["product_code"], 0.0)
            expected_amount = round(p["promised_qty"] * list_price, 2)

        out.append(
            {
                **p,
                "expected_amount": expected_amount,
            }
        )
    return out


def build_route_agg(routes: list) -> dict:
    agg = defaultdict(lambda: {
        "planned_visits": 0,
        "actual_visits": 0,
        "km_travelled": 0.0,
        "travel_expense": 0.0,
        "road_toll": 0.0,
        "route_count": 0,
    })

    for r in routes:
        key = (r["seller_code"], r["visit_date"])
        if not key[0] or not key[1]:
            continue
        a = agg[key]
        a["planned_visits"] += r["planned_visits"]
        a["actual_visits"] += r["actual_visits"]
        a["km_travelled"] += r["km_travelled"]
        a["travel_expense"] += r["travel_expense"]
        a["road_toll"] += r["road_toll"]
        a["route_count"] += 1

    return agg


def build_fuel_agg(fuel: list) -> dict:
    agg = defaultdict(lambda: {
        "fuel_liters": 0.0,
        "fuel_cost": 0.0,
        "toll_cost": 0.0,
        "hotel_cost": 0.0,
        "meal_cost": 0.0,
        "misc_cost": 0.0,
        "fuel_count": 0,
    })

    for f in fuel:
        key = (f["seller_code"], f["expense_date"])
        if not key[0] or not key[1]:
            continue
        a = agg[key]
        a["fuel_liters"] += f["fuel_liters"]
        a["fuel_cost"] += f["fuel_cost"]
        a["toll_cost"] += f["toll_cost"]
        a["hotel_cost"] += f["hotel_cost"]
        a["meal_cost"] += f["meal_cost"]
        a["misc_cost"] += f["misc_cost"]
        a["fuel_count"] += 1

    return agg


def transform_fact(
    orders: list,
    promises: list,
    routes: list,
    fuel: list,
    dim_date: list,
    dim_seller: list,
    dim_customer: list,
    dim_product: list,
    dim_geo: list,
) -> list:
    date_idx = {r["date_complete"]: r["id_dim_date"] for r in dim_date}
    seller_idx = {r["seller_code"]: r["id_dim_seller"] for r in dim_seller}
    customer_idx = {r["customer_code"]: r["id_dim_customer"] for r in dim_customer}
    product_idx = {r["product_code"]: r["id_dim_product"] for r in dim_product}

    customer_city_idx = {}
    for c in dim_customer:
        customer_city_idx[c["customer_code"]] = norm_city(c.get("city"))
    geo_idx = {r["city"]: r["id_dim_geo"] for r in dim_geo}

    route_agg = build_route_agg(routes)
    fuel_agg = build_fuel_agg(fuel)

    # Index promesses au grain date+seller+customer+product
    promise_idx = {}
    for p in promises:
        k = (p["promise_date"], p["seller_code"], p["customer_code"], p["product_code"])
        if k not in promise_idx:
            promise_idx[k] = {
                "promised_qty": 0,
                "expected_amount": 0.0,
                "probability_pct": 0.0,
                "status": p["status"],
                "count": 0,
            }
        v = promise_idx[k]
        v["promised_qty"] += p["promised_qty"]
        v["expected_amount"] += p["expected_amount"]
        v["probability_pct"] += p["probability_pct"]
        v["count"] += 1

    facts = []

    # Lignes ventes
    for o in orders:
        date_key = o["order_date"]
        seller_code = o["seller_code"]
        customer_code = o["customer_code"]
        product_code = o["product_code"]

        id_date = date_idx.get(date_key)
        id_seller = seller_idx.get(seller_code)
        id_customer = customer_idx.get(customer_code)
        id_product = product_idx.get(product_code)

        city = customer_city_idx.get(customer_code, "INCONNUE")
        id_geo = geo_idx.get(city, geo_idx.get("INCONNUE"))

        if not all([id_date, id_seller, id_customer, id_product, id_geo]):
            continue

        gross = round(o["quantity"] * o["unit_price"], 2)
        net_sales_amount = round(gross * (1.0 - o["discount_pct"]), 2)

        promise_key = (date_key, seller_code, customer_code, product_code)
        p = promise_idx.get(promise_key, None)
        promised_qty = p["promised_qty"] if p else 0
        expected_amount = round(p["expected_amount"], 2) if p else 0.0
        probability_pct = round((p["probability_pct"] / p["count"]), 2) if p else 0.0
        promise_status = p["status"] if p else None

        facts.append(
            {
                "id_dim_date": id_date,
                "id_dim_seller": id_seller,
                "id_dim_customer": id_customer,
                "id_dim_product": id_product,
                "id_dim_geo": id_geo,
                "source_type": "ORDER",
                "order_id": o["order_id"],
                "promise_id": None,
                "quantity": o["quantity"],
                "unit_price": round(o["unit_price"], 2),
                "discount_pct": round(o["discount_pct"], 4),
                "gross_sales_amount": gross,
                "net_sales_amount": net_sales_amount,
                "promised_qty": promised_qty,
                "expected_amount": expected_amount,
                "probability_pct": probability_pct,
                "km_travelled": 0.0,
                "travel_expense": 0.0,
                "road_toll": 0.0,
                "fuel_liters": 0.0,
                "fuel_cost": 0.0,
                "hotel_cost": 0.0,
                "meal_cost": 0.0,
                "misc_cost": 0.0,
                "order_status": o["order_status"],
                "promise_status": promise_status,
                "activity_status": o["order_status"],
                "route_count": 0,
                "fuel_count": 0,
            }
        )

    # Lignes promesses sans vente le même jour
    existing_keys = {
        (f["id_dim_date"], f["id_dim_seller"], f["id_dim_customer"], f["id_dim_product"]) for f in facts
    }
    for p in promises:
        id_date = date_idx.get(p["promise_date"])
        id_seller = seller_idx.get(p["seller_code"])
        id_customer = customer_idx.get(p["customer_code"])
        id_product = product_idx.get(p["product_code"])
        city = customer_city_idx.get(p["customer_code"], "INCONNUE")
        id_geo = geo_idx.get(city, geo_idx.get("INCONNUE"))

        if not all([id_date, id_seller, id_customer, id_product, id_geo]):
            continue

        key = (id_date, id_seller, id_customer, id_product)
        if key in existing_keys:
            continue

        facts.append(
            {
                "id_dim_date": id_date,
                "id_dim_seller": id_seller,
                "id_dim_customer": id_customer,
                "id_dim_product": id_product,
                "id_dim_geo": id_geo,
                "source_type": "PROMISE",
                "order_id": None,
                "promise_id": p["promise_id"],
                "quantity": 0,
                "unit_price": 0.0,
                "discount_pct": 0.0,
                "gross_sales_amount": 0.0,
                "net_sales_amount": 0.0,
                "promised_qty": p["promised_qty"],
                "expected_amount": round(p["expected_amount"], 2),
                "probability_pct": round(p["probability_pct"], 2),
                "km_travelled": 0.0,
                "travel_expense": 0.0,
                "road_toll": 0.0,
                "fuel_liters": 0.0,
                "fuel_cost": 0.0,
                "hotel_cost": 0.0,
                "meal_cost": 0.0,
                "misc_cost": 0.0,
                "order_status": None,
                "promise_status": p["status"],
                "activity_status": p["status"],
                "route_count": 0,
                "fuel_count": 0,
            }
        )

    # Allocation des coûts route+fuel sur la première ligne de chaque vendeur+date
    grouped_indices = defaultdict(list)
    for i, f in enumerate(facts):
        grouped_indices[(f["id_dim_seller"], f["id_dim_date"])].append(i)

    seller_rev = {v: k for k, v in seller_idx.items()}
    date_rev = {v: k for k, v in date_idx.items()}

    for (id_seller, id_date), indexes in grouped_indices.items():
        seller_code = seller_rev[id_seller]
        d_str = date_rev[id_date]
        r = route_agg.get((seller_code, d_str), None)
        fu = fuel_agg.get((seller_code, d_str), None)

        if not r and not fu:
            continue

        first = indexes[0]
        if r:
            facts[first]["km_travelled"] = round(r["km_travelled"], 2)
            facts[first]["travel_expense"] = round(r["travel_expense"], 2)
            facts[first]["road_toll"] = round(r["road_toll"], 2)
            facts[first]["route_count"] = r["route_count"]
        if fu:
            facts[first]["fuel_liters"] = round(fu["fuel_liters"], 2)
            facts[first]["fuel_cost"] = round(fu["fuel_cost"], 2)
            facts[first]["hotel_cost"] = round(fu["hotel_cost"], 2)
            facts[first]["meal_cost"] = round(fu["meal_cost"], 2)
            facts[first]["misc_cost"] = round(fu["misc_cost"], 2)
            facts[first]["road_toll"] = round(facts[first]["road_toll"] + fu["toll_cost"], 2)
            facts[first]["fuel_count"] = fu["fuel_count"]

    return facts


# ══════════════════════════════════════════════════════════════════════════════
# ÉTAPE 3 — CHARGEMENT PostgreSQL (génération SQL)
# ══════════════════════════════════════════════════════════════════════════════
DDL_POSTGRES = """
DROP TABLE IF EXISTS fact_sales_activity CASCADE;
DROP TABLE IF EXISTS dim_date CASCADE;
DROP TABLE IF EXISTS dim_seller CASCADE;
DROP TABLE IF EXISTS dim_customer CASCADE;
DROP TABLE IF EXISTS dim_product CASCADE;
DROP TABLE IF EXISTS dim_geo CASCADE;

CREATE TABLE dim_date (
    id_dim_date      SERIAL PRIMARY KEY,
    date_complete    DATE NOT NULL,
    annee            SMALLINT NOT NULL,
    trimestre        SMALLINT NOT NULL,
    mois             SMALLINT NOT NULL,
    lib_mois         VARCHAR(20) NOT NULL,
    semaine          SMALLINT NOT NULL,
    jour             SMALLINT NOT NULL,
    lib_jour         VARCHAR(20) NOT NULL,
    est_weekend      BOOLEAN NOT NULL DEFAULT FALSE
);

CREATE TABLE dim_seller (
    id_dim_seller    SERIAL PRIMARY KEY,
    seller_code      VARCHAR(10) UNIQUE,
    first_name       VARCHAR(60),
    last_name        VARCHAR(60),
    full_name        VARCHAR(130),
    email            VARCHAR(120),
    salary           NUMERIC(12,2),
    hire_date        DATE,
    home_country     VARCHAR(60),
    home_region      VARCHAR(80),
    home_city        VARCHAR(80),
    manager_code     VARCHAR(10)
);

CREATE TABLE dim_customer (
    id_dim_customer  SERIAL PRIMARY KEY,
    customer_code    VARCHAR(10) UNIQUE,
    customer_name    VARCHAR(150),
    sector           VARCHAR(50),
    country          VARCHAR(60),
    region_name      VARCHAR(80),
    city             VARCHAR(80),
    postal_code      VARCHAR(20),
    created_at       DATE
);

CREATE TABLE dim_product (
    id_dim_product   SERIAL PRIMARY KEY,
    product_code     VARCHAR(10) UNIQUE,
    product_name     VARCHAR(120),
    category_name    VARCHAR(60),
    range_name       VARCHAR(50),
    list_price       NUMERIC(12,2),
    active_flag      BOOLEAN,
    launch_date      DATE
);

CREATE TABLE dim_geo (
    id_dim_geo       SERIAL PRIMARY KEY,
    city             VARCHAR(80),
    region_name      VARCHAR(80),
    country          VARCHAR(60)
);

CREATE TABLE fact_sales_activity (
    id_fact              SERIAL PRIMARY KEY,
    id_dim_date          INT NOT NULL REFERENCES dim_date(id_dim_date),
    id_dim_seller        INT NOT NULL REFERENCES dim_seller(id_dim_seller),
    id_dim_customer      INT NOT NULL REFERENCES dim_customer(id_dim_customer),
    id_dim_product       INT NOT NULL REFERENCES dim_product(id_dim_product),
    id_dim_geo           INT NOT NULL REFERENCES dim_geo(id_dim_geo),
    source_type          VARCHAR(15),
    order_id             INT,
    promise_id           VARCHAR(20),
    quantity             INT,
    unit_price           NUMERIC(12,2),
    discount_pct         NUMERIC(8,4),
    gross_sales_amount   NUMERIC(14,2),
    net_sales_amount     NUMERIC(14,2),
    promised_qty         INT,
    expected_amount      NUMERIC(14,2),
    probability_pct      NUMERIC(6,2),
    km_travelled         NUMERIC(12,2),
    travel_expense       NUMERIC(12,2),
    road_toll            NUMERIC(12,2),
    fuel_liters          NUMERIC(12,2),
    fuel_cost            NUMERIC(12,2),
    hotel_cost           NUMERIC(12,2),
    meal_cost            NUMERIC(12,2),
    misc_cost            NUMERIC(12,2),
    order_status         VARCHAR(20),
    promise_status       VARCHAR(20),
    activity_status      VARCHAR(20),
    route_count          INT,
    fuel_count           INT
);

CREATE INDEX idx_fact_date ON fact_sales_activity(id_dim_date);
CREATE INDEX idx_fact_seller ON fact_sales_activity(id_dim_seller);
CREATE INDEX idx_fact_customer ON fact_sales_activity(id_dim_customer);
CREATE INDEX idx_fact_product ON fact_sales_activity(id_dim_product);
CREATE INDEX idx_fact_geo ON fact_sales_activity(id_dim_geo);
"""


def sql_nullable(value):
    if value is None or value == "":
        return "NULL"
    return f"'{esc_sql(str(value))}'"


def sql_bool(value: bool):
    return "TRUE" if value else "FALSE"


def generate_load_sql(dim_date, dim_seller, dim_customer, dim_product, dim_geo, facts) -> str:
    lines = [DDL_POSTGRES]

    lines.append("\n-- DIM_DATE")
    for r in dim_date:
        lines.append(
            "INSERT INTO dim_date(id_dim_date,date_complete,annee,trimestre,mois,lib_mois,semaine,jour,lib_jour,est_weekend) "
            f"VALUES ({r['id_dim_date']},'{r['date_complete']}',{r['annee']},{r['trimestre']},{r['mois']},"
            f"'{esc_sql(r['lib_mois'])}',{r['semaine']},{r['jour']},'{esc_sql(r['lib_jour'])}',{sql_bool(r['est_weekend'])});"
        )

    lines.append("\n-- DIM_SELLER")
    for r in dim_seller:
        lines.append(
            "INSERT INTO dim_seller(id_dim_seller,seller_code,first_name,last_name,full_name,email,salary,hire_date,home_country,home_region,home_city,manager_code) "
            f"VALUES ({r['id_dim_seller']},'{esc_sql(r['seller_code'])}','{esc_sql(r['first_name'])}','{esc_sql(r['last_name'])}',"
            f"'{esc_sql(r['full_name'])}',{sql_nullable(r['email'])},{round(r['salary'], 2)},{sql_nullable(r['hire_date'])},"
            f"{sql_nullable(r['home_country'])},{sql_nullable(r['home_region'])},{sql_nullable(r['home_city'])},{sql_nullable(r['manager_code'])});"
        )

    lines.append("\n-- DIM_CUSTOMER")
    for r in dim_customer:
        lines.append(
            "INSERT INTO dim_customer(id_dim_customer,customer_code,customer_name,sector,country,region_name,city,postal_code,created_at) "
            f"VALUES ({r['id_dim_customer']},'{esc_sql(r['customer_code'])}','{esc_sql(r['customer_name'])}',"
            f"{sql_nullable(r['sector'])},{sql_nullable(r['country'])},{sql_nullable(r['region_name'])},"
            f"{sql_nullable(r['city'])},{sql_nullable(r['postal_code'])},{sql_nullable(r['created_at'])});"
        )

    lines.append("\n-- DIM_PRODUCT")
    for r in dim_product:
        lines.append(
            "INSERT INTO dim_product(id_dim_product,product_code,product_name,category_name,range_name,list_price,active_flag,launch_date) "
            f"VALUES ({r['id_dim_product']},'{esc_sql(r['product_code'])}','{esc_sql(r['product_name'])}',"
            f"{sql_nullable(r['category_name'])},{sql_nullable(r['range_name'])},{round(r['list_price'], 2)},"
            f"{sql_bool(r['active_flag'])},{sql_nullable(r['launch_date'])});"
        )

    lines.append("\n-- DIM_GEO")
    for r in dim_geo:
        lines.append(
            "INSERT INTO dim_geo(id_dim_geo,city,region_name,country) "
            f"VALUES ({r['id_dim_geo']},'{esc_sql(r['city'])}','{esc_sql(r['region_name'])}','{esc_sql(r['country'])}');"
        )

    lines.append("\n-- FACT_SALES_ACTIVITY")
    for f in facts:
        lines.append(
            "INSERT INTO fact_sales_activity("
            "id_dim_date,id_dim_seller,id_dim_customer,id_dim_product,id_dim_geo,"
            "source_type,order_id,promise_id,quantity,unit_price,discount_pct,gross_sales_amount,net_sales_amount,"
            "promised_qty,expected_amount,probability_pct,km_travelled,travel_expense,road_toll,"
            "fuel_liters,fuel_cost,hotel_cost,meal_cost,misc_cost,order_status,promise_status,activity_status,route_count,fuel_count"
            ") VALUES ("
            f"{f['id_dim_date']},{f['id_dim_seller']},{f['id_dim_customer']},{f['id_dim_product']},{f['id_dim_geo']},"
            f"'{esc_sql(f['source_type'])}',{f['order_id'] if f['order_id'] is not None else 'NULL'},{sql_nullable(f['promise_id'])},"
            f"{f['quantity']},{round(f['unit_price'], 2)},{round(f['discount_pct'], 4)},{round(f['gross_sales_amount'], 2)},"
            f"{round(f['net_sales_amount'], 2)},{f['promised_qty']},{round(f['expected_amount'], 2)},{round(f['probability_pct'], 2)},"
            f"{round(f['km_travelled'], 2)},{round(f['travel_expense'], 2)},{round(f['road_toll'], 2)},"
            f"{round(f['fuel_liters'], 2)},{round(f['fuel_cost'], 2)},{round(f['hotel_cost'], 2)},{round(f['meal_cost'], 2)},"
            f"{round(f['misc_cost'], 2)},{sql_nullable(f['order_status'])},{sql_nullable(f['promise_status'])},"
            f"{sql_nullable(f['activity_status'])},{f['route_count']},{f['fuel_count']});"
        )

    lines.append(
        """
-- Requêtes analytiques de validation

-- 1) Chiffre d'affaires net par vendeur et par mois
SELECT s.full_name, d.annee, d.mois,
       SUM(f.net_sales_amount) AS ca_net
FROM fact_sales_activity f
JOIN dim_seller s ON f.id_dim_seller = s.id_dim_seller
JOIN dim_date d   ON f.id_dim_date = d.id_dim_date
GROUP BY s.full_name, d.annee, d.mois
ORDER BY d.annee, d.mois, ca_net DESC;

-- 2) Kilomètres parcourus par région
SELECT g.region_name,
       SUM(f.km_travelled) AS km_total
FROM fact_sales_activity f
JOIN dim_geo g ON f.id_dim_geo = g.id_dim_geo
GROUP BY g.region_name
ORDER BY km_total DESC;

-- 3) Efficacité vendeurs : km vs CA
SELECT s.full_name,
       SUM(f.km_travelled) AS km_total,
       SUM(f.net_sales_amount) AS ca_net,
       CASE WHEN SUM(f.km_travelled) = 0 THEN NULL
            ELSE ROUND(SUM(f.net_sales_amount) / SUM(f.km_travelled), 2)
       END AS ca_par_km
FROM fact_sales_activity f
JOIN dim_seller s ON f.id_dim_seller = s.id_dim_seller
GROUP BY s.full_name
ORDER BY ca_par_km ASC NULLS LAST;

-- 4) Clients les plus coûteux à servir
SELECT c.customer_name,
       SUM(f.travel_expense + f.road_toll + f.fuel_cost + f.hotel_cost + f.meal_cost + f.misc_cost) AS cout_terrain,
       SUM(f.net_sales_amount) AS ca_net
FROM fact_sales_activity f
JOIN dim_customer c ON f.id_dim_customer = c.id_dim_customer
GROUP BY c.customer_name
ORDER BY cout_terrain DESC;

-- 5) Taux de transformation promesses -> ventes
SELECT d.annee, d.mois,
       SUM(CASE WHEN f.promised_qty > 0 THEN 1 ELSE 0 END) AS nb_promesses,
       SUM(CASE WHEN f.net_sales_amount > 0 THEN 1 ELSE 0 END) AS nb_ventes,
       ROUND(
           100.0 * SUM(CASE WHEN f.net_sales_amount > 0 THEN 1 ELSE 0 END)
           / NULLIF(SUM(CASE WHEN f.promised_qty > 0 THEN 1 ELSE 0 END), 0),
           2
       ) AS taux_transformation_pct
FROM fact_sales_activity f
JOIN dim_date d ON f.id_dim_date = d.id_dim_date
GROUP BY d.annee, d.mois
ORDER BY d.annee, d.mois;
"""
    )

    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════
def run_etl() -> bool:
    print(f"\n{B}{'═' * 60}{E}")
    print(f"{B}  ETL DataWarehouse — Activité commerciale imprimantes{E}")
    print(f"{B}{'═' * 60}{E}\n")

    # E - Extraction
    print(f"{Y}[E] EXTRACTION{E}")

    info("MySQL SQL -> vendeurs, clients, produits, ventes")
    sellers, customers, products, orders = extract_mysql_sql(SQL_FILE)
    ok(f"{len(sellers)} vendeurs | {len(customers)} clients | {len(products)} produits | {len(orders)} ventes")

    info("TXT route_logs -> déplacements")
    routes = extract_route_logs(TXT_FILE)
    ok(f"{len(routes)} lignes route")

    info("Excel promesses de vente")
    promises = extract_excel_promises(EXCEL_FILE)
    ok(f"{len(promises)} promesses")

    info("JSON carburant / frais")
    fuel = extract_fuel_json(JSON_FILE)
    ok(f"{len(fuel)} lignes fuel")

    # T - Transformation
    print(f"\n{Y}[T] TRANSFORMATION{E}")

    product_price_index = {norm_code(p["product_code"]): parse_decimal(p["list_price"]) for p in products}
    promises = transform_promises(promises, product_price_index)

    info("Dimensions")
    dim_date = transform_dim_date(orders, promises, routes, fuel)
    dim_seller = transform_dim_seller(sellers, orders, promises, routes, fuel)
    dim_customer = transform_dim_customer(customers, orders, promises, routes)
    dim_product = transform_dim_product(products, orders, promises)
    dim_geo = transform_dim_geo(customers, routes, fuel, sellers)
    ok(
        f"date={len(dim_date)} | seller={len(dim_seller)} | customer={len(dim_customer)} | "
        f"product={len(dim_product)} | geo={len(dim_geo)}"
    )

    info("Table de faits")
    facts = transform_fact(
        orders,
        promises,
        routes,
        fuel,
        dim_date,
        dim_seller,
        dim_customer,
        dim_product,
        dim_geo,
    )
    ok(f"{len(facts)} lignes fact_sales_activity")

    # L - Load (SQL)
    print(f"\n{Y}[L] CHARGEMENT{E}")
    info("Génération du script SQL PostgreSQL")
    sql_text = generate_load_sql(dim_date, dim_seller, dim_customer, dim_product, dim_geo, facts)
    OUT_SQL.write_text(sql_text, encoding="utf-8")
    ok(f"Script généré -> {OUT_SQL.name}")

    print(f"\n{B}{'═' * 60}{E}")
    print(f"{G}  ETL TERMINE AVEC SUCCES{E}")
    print(f"{B}{'═' * 60}{E}")
    print(
        f"""
  Resume du chargement :
    DIM_DATE        : {len(dim_date):>6}
    DIM_SELLER      : {len(dim_seller):>6}
    DIM_CUSTOMER    : {len(dim_customer):>6}
    DIM_PRODUCT     : {len(dim_product):>6}
    DIM_GEO         : {len(dim_geo):>6}
    FACT            : {len(facts):>6}

  Fichier genere : {OUT_SQL.name}
"""
    )

    return True


def parse_args():
    parser = argparse.ArgumentParser(description="ETL TP Datawarehouse")
    return parser.parse_args()


if __name__ == "__main__":
    parse_args()
    success = run_etl()
    sys.exit(0 if success else 1)
